import csv
import io
from typing import Type


class Reader:
    def __init__(self, *, file: Type[io.BytesIO], encoding: str = 'cp932',
                 table_start_from: int = 0, **kwargs):
        self.file = file
        self.encoding = encoding
        self.table_start_from = table_start_from
        self.kwargs = kwargs


class CsvMixin:
    def get_table(self) -> list:
        csv_file = io.TextIOWrapper(
            self.file.file, encoding=self.encoding, **self.kwargs
        )

        return list(csv.reader(csv_file, delimiter=self.delimiter)
                    )[self.table_start_from:]


class CsvReader(CsvMixin, Reader):
    delimiter = ','


class TsvReader(CsvMixin, Reader):
    delimiter = '\t'


class XlsReader(Reader):
    def get_table(self, sheet_index: int = 0) -> list:
        import xlrd

        wb = xlrd.open_workbook(file_contents=self.file.read(), **self.kwargs)
        sheet = wb.sheets()[sheet_index]
        return [['' if value is None else value
                 for value in sheet.row_values(n)]
                for n in range(self.table_start_from, sheet.nrows)]


class XlsxReader(Reader):
    def get_table(self, sheet_index: int = 0) -> list:
        from openpyxl import load_workbook

        wb = load_workbook(self.file, data_only=True)
        sheet = wb.worksheets[sheet_index]
        table = [['' if cell.value is None else cell.value for cell in row]
                 for row in sheet.iter_rows(min_row=self.table_start_from + 1)]
        return table


class ExcelReader:
    def __init__(self, *, file: Type[io.BytesIO], **kwargs):
        if file.name.lower().endswith('.xls'):
            self.reader = XlsReader(file=file, **kwargs)
        elif file.name.lower().endswith('.xlsx'):
            self.reader = XlsxReader(file=file, **kwargs)

    def get_table(self, sheet_index: int = 0) -> list:
        return self.reader.get_table(sheet_index)
