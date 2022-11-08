import csv
import io

import urllib
from django.http import HttpResponse
from typing import Optional


class Writer:
    content_type = 'text/csv'
    extension = 'csv'

    def __init__(self, filename: str, encoding: str = 'utf-8') -> None:
        self.filename = filename if filename.endswith(
            self.extension) else f'{filename}.{self.extension}'
        if len(self.filename.split('.')) != 2:
            raise ValueError(f'{filename} may have unexpected extension. '
                             f'`.{self.extension}` is expected.')
        self.encoding = encoding

    def make_response(self, **kwargs) -> HttpResponse:
        return self._response()

    def write_down(self, table: list, **kwargs) -> None:
        pass

    def _response(self, **kwargs) -> HttpResponse:
        res = HttpResponse(content_type=self.content_type)
        filename = urllib.parse.quote(self.filename)
        res['Content-Disposition'] = f'attachment;filename="{filename}"'
        return res


class CsvMixin:
    delimiter = None

    def __init__(self, *args, **kwargs):
        self.table = []
        super().__init__(*args, **kwargs)

    def write_down(self, table: list, separator: list = None) -> None:
        if separator:
            self.table.extend(separator)

        self.table.extend(table)

    def make_response(self, **kwargs):
        with io.StringIO() as sio:
            w = csv.writer(sio, delimiter=self.delimiter)
            w.writerows(self.table)
            res = self._response()
            res.write(sio.getvalue().encode(self.encoding, errors='ignore'))

        return res


class CsvWriter(CsvMixin, Writer):
    delimiter = ','
    extension = 'csv'
    content_type = 'text/csv'


class TsvWriter(CsvMixin, Writer):
    delimiter = '\t'
    extension = 'tsv'
    content_type = 'text/tsv'


class ExcelMixin:
    content_type = 'application/vnd.ms-excel'

    def __init__(self, *args, default_sheet_name: str = 'sheet', **kwargs):
        self.default_sheet_name = default_sheet_name
        super().__init__(*args, **kwargs)

    def _get_next_sheet_name(self) -> str:
        return f'{self.default_sheet_name} {len(self.get_sheet_names())}'


class XlsWriter(ExcelMixin, Writer):
    extension = 'xls'

    def __init__(self, *args, **kwargs):
        import xlwt
        self.wb = xlwt.Workbook()
        self.sheet_names = []
        super().__init__(*args, **kwargs)

    def get_sheet_names(self) -> list:
        return self.sheet_names

    def write_down(self, table: list, sheet_name: Optional[str] = None):
        """
        write down to work sheet.
        """
        if sheet_name is None:
            sheet_name = self._get_next_sheet_name()

        ws = self.wb.add_sheet(sheet_name)
        for y, row in enumerate(table):
            for x, col in enumerate(row):
                ws.write(y, x, col)
        self.sheet_names.append(sheet_name)

    def make_response(self, **kwargs):
        res = self._response()
        self.wb.save(res)
        return res


class XlsxWriter(ExcelMixin, Writer):
    extension = 'xlsx'

    def __init__(self, *args, **kwargs):
        from openpyxl import Workbook
        self.wb = Workbook()
        # delete default sheet
        self.wb.remove(self.wb.active)
        super().__init__(*args, **kwargs)

    def get_sheet_names(self) -> list:
        return self.wb.sheetnames

    def write_down(self, table: list, sheet_name: Optional[str] = None) -> None:
        """
        write down to work sheet.
        """
        if sheet_name is None:
            sheet_name = self._get_next_sheet_name()

        ws = self.wb.create_sheet(sheet_name)
        for y, row in enumerate(table, 1):
            for x, col in enumerate(row, 1):
                ws.cell(y, x).value = col

    def make_response(self, **kwargs):
        res = self._response()
        self.wb.save(res)
        return res
