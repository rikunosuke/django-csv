import csv
import io
from django.core.files import File

from datetime import datetime
from typing import Union, TextIO, Optional


class Reader:
    convert_to_stringio: bool = False
    def __init__(self, *, file: Union[TextIO, File], encoding: str = 'utf-8',
                 table_starts_from: int = 0, **kwargs):
        if isinstance(file, File):
            f = file.file
        else:
            f = file
        if self.convert_to_stringio and isinstance(f, io.BytesIO):
            f = io.TextIOWrapper(f, encoding=encoding)

        self.file: io.BytesIO | io.StringIO = f
        self.encoding = encoding
        self.table_starts_from = table_starts_from
        self.kwargs = kwargs


class CsvBase(Reader):
    delimiter = None
    convert_to_stringio = True

    def get_table(self, table_starts_from: Optional[int] = None) -> list:
        table_starts_from = self.table_starts_from \
            if table_starts_from is None else table_starts_from

        return list(csv.reader(self.file, delimiter=self.delimiter)
                    )[table_starts_from:]


class CsvReader(CsvBase):
    delimiter = ','


class TsvReader(CsvBase):
    delimiter = '\t'


class ExcelBase(Reader):
    def __init__(self, date_format: str = None, datetime_format: str = None,
                 **kwargs):
        self.date_format = date_format or '%Y-%m-%d'
        self.datetime_format = datetime_format or '%Y-%m-%d %H:%M:%s'
        super().__init__(**kwargs)


class XlsReader(ExcelBase):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        import xlrd
        self.wb = xlrd.open_workbook(
            file_contents=self.file.read(), **self.kwargs)
        self.xldate_as_tuple = xlrd.xldate_as_tuple

    def get_table(self, sheet_index: int = 0,
                  table_starts_from: Optional[int] = None
                  ) -> list:
        sheet = self.wb.sheets()[sheet_index]
        table_starts_from = self.table_starts_from \
            if table_starts_from is None else table_starts_from

        def __get_value(_cell) -> str:
            if _cell.value is None:
                return ''

            if _cell.ctype == 3:
                _d = datetime(
                    *self.xldate_as_tuple(_cell.value, self.wb.datemode))
                if all([
                    _d.hour == 0, _d.minute == 0, _d.second == 0,
                    _d.microsecond == 0]
                ):
                    return _d.date().strftime(self.date_format)
                else:
                    return _d.strftime(self.datetime_format)

            elif _cell.ctype == 2:
                if _cell.value.is_integer():
                    return str(int(_cell.value))
                else:
                    return str(_cell.value)

            return str(_cell.value)

        return [
            [__get_value(sheet.cell(y, x)) for x in range(sheet.ncols)]
            for y in range(table_starts_from, sheet.nrows)
        ]

    def get_sheet_names(self) -> list:
        return self.wb.sheetnames


class XlsxReader(ExcelBase):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        from openpyxl import load_workbook
        self.wb = load_workbook(self.file, data_only=True)

    def get_table(self, sheet_index: int = 0,
                  table_starts_from: Optional[int] = None) -> list:
        table_starts_from = self.table_starts_from \
            if table_starts_from is None else table_starts_from
        sheet = self.wb.worksheets[sheet_index]

        def __get_value(_cell) -> str:
            _val = _cell.value
            if _val is None:
                return ''

            if isinstance(_val, str):
                return _val

            if isinstance(_val, (int, float)):
                _val = str(_val)
                if _val.endswith('.0'):
                    _val = _val[:-2]
                return _val

            if isinstance(_val, datetime):
                if not any([_val.hour, _val.minute, _val.microsecond]):
                    return _val.date().strftime(self.date_format)
                return _val.strftime(self.datetime_format)
            return str(_val)

        return [
            [__get_value(cell) for cell in row]
            for row in sheet.iter_rows(min_row=table_starts_from + 1)
        ]

    def get_sheet_names(self) -> list:
        return self.wb.sheetnames
